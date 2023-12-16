from openpyxl import load_workbook

class Fatura:
    def __init__(self, caminho_arquivo):
        self.caminho = caminho_arquivo

    def exibir_fatura(self): 
        data = input('Data: ')
        descricao = input('Descrição: ')
        valor = float(input('Valor: '))

        # Carregar o arquivo Excel existente ou criar um novo
        try:
            wb = load_workbook(self.caminho)
        except FileNotFoundError:
            wb = Workbook()

        # Selecionar a planilha 'Faturas' ou criar uma nova
        if 'Faturas' in wb.sheetnames:
            ws = wb['Faturas']
        else:
            ws = wb.create_sheet('Faturas')

        # Adicionar novos dados à planilha
        nova_linha = [data, descricao, valor]
        ws.append(nova_linha)

        # Salvar o arquivo Excel
        wb.save(self.caminho)

# Caminho para o arquivo Excel
caminho = r'C:\Users\sivei\OneDrive\Documentos\Projeto-Computacional\bancodedados.xlsx'

# Criar uma instância da classe Fatura passando o caminho do arquivo
fatura = Fatura(caminho)
fatura.exibir_fatura()
